from flask import Flask, request, jsonify, send_file, session, send_from_directory
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
import os
from datetime import datetime
import json
import sqlite3

print("🚀 APP.PY YÜKLENDI - VERİTABANI İLE!")

# Veritabanı başlatma
def init_db():
    try:
        conn = sqlite3.connect('nobet.db')
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS published_tables
                     (id INTEGER PRIMARY KEY, data TEXT, created_at TIMESTAMP)''')
        c.execute('''CREATE TABLE IF NOT EXISTS physicians_data
                     (id INTEGER PRIMARY KEY, data TEXT, created_at TIMESTAMP)''')
        c.execute('''CREATE TABLE IF NOT EXISTS app_settings
                     (id INTEGER PRIMARY KEY, data TEXT, created_at TIMESTAMP)''')
        conn.commit()
        conn.close()
        print("✅ Veritabanı hazırlandı")
    except Exception as e:
        print(f"⚠️ DB hata: {e}")

init_db()

app = Flask(__name__, static_folder='.')
app.secret_key = 'nobet-sistemi-secret-key-2026'  # Session için secret key

# CORS - Basit ve açık
CORS(app, resources={r"/api/*": {"origins": "*"}}, supports_credentials=True)

# Session ayarları
app.config['SESSION_COOKIE_SAMESITE'] = 'None'  # Cross-domain için None
app.config['SESSION_COOKIE_SECURE'] = True  # HTTPS gerektirir
app.config['SESSION_COOKIE_HTTPONLY'] = True

# Manuel CORS header'ları ekle
@app.after_request
def after_request(response):
    response.headers['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
    response.headers['Access-Control-Allow-Credentials'] = 'true'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    return response

# Kullanıcılar (gerçek uygulamada veritabanında olmalı)
USERS = {
    'admin': {'password': 'admin123', 'role': 'admin'},
    'pratisyen': {'password': 'prat123', 'role': 'user'}
}

# Yayınlanan tablo (gerçek uygulamada veritabanında olmalı)
published_data = None

# Frontend routes
@app.route('/')
def index():
    """Ana sayfa - login'e yönlendir"""
    return send_from_directory('.', 'index.html')

@app.route('/<path:path>')
def static_files(path):
    """Static dosyaları servis et"""
    try:
        return send_from_directory('.', path)
    except:
        return "Dosya bulunamadı", 404

@app.route('/health', methods=['GET'])
def health_check():
    """Sunucu sağlık kontrolü"""
    return jsonify({'status': 'ok', 'message': 'Backend çalışıyor!'})

@app.route('/api/login', methods=['POST'])
def login():
    """Kullanıcı girişi"""
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    print(f"🔐 Login denemesi: {username}")
    
    if username in USERS and USERS[username]['password'] == password:
        session['username'] = username
        session['role'] = USERS[username]['role']
        print(f"✅ Login başarılı: {username} - Role: {session['role']}")
        return jsonify({
            'success': True,
            'username': username,
            'role': USERS[username]['role']
        })
    
    print(f"❌ Login başarısız: {username}")
    return jsonify({'success': False, 'message': 'Kullanıcı adı veya şifre hatalı'}), 401

@app.route('/api/logout', methods=['POST'])
def logout():
    """Çıkış yap"""
    print(f"🚪 Logout: {session.get('username', 'Bilinmeyen')}")
    session.clear()
    return jsonify({'success': True})

@app.route('/api/check-auth', methods=['GET'])
def check_auth():
    """Oturum kontrolü"""
    if 'username' in session:
        print(f"✅ Auth OK: {session['username']} - {session['role']}")
        return jsonify({
            'authenticated': True,
            'username': session['username'],
            'role': session['role']
        })
    print("❌ Auth yok")
    return jsonify({'authenticated': False})

@app.route('/api/publish', methods=['POST'])
def publish():
    """Tabloyu yayınla (sadece admin)"""
    
    print(f"📢 Publish isteği - Session: {session}")
    
    if session.get('role') != 'admin':
        print(f"❌ Yetki yok! Session role: {session.get('role')}")
        return jsonify({'success': False, 'message': 'Yetkiniz yok'}), 403
    
    try:
        data = request.json
        conn = sqlite3.connect('nobet.db')
        c = conn.cursor()
        
        # Son kaydı sil, yenisini ekle (tüm uygulamada tek tablo)
        c.execute('DELETE FROM published_tables')
        c.execute('INSERT INTO published_tables (data, created_at) VALUES (?, ?)',
                  (json.dumps(data), datetime.now()))
        conn.commit()
        conn.close()
        
        print(f"✅ Tablo yayınlandı ve veritabanına kaydedildi!")
        return jsonify({'success': True, 'message': 'Tablo yayınlandı'})
    except Exception as e:
        print(f"❌ Publish hatası: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/published', methods=['GET'])
def get_published():
    """Yayınlanan tabloyu getir"""
    print(f"👁️ Published isteği - Auth: {session.get('username', 'Yok')}")
    
    try:
        conn = sqlite3.connect('nobet.db')
        c = conn.cursor()
        c.execute('SELECT data FROM published_tables ORDER BY created_at DESC LIMIT 1')
        result = c.fetchone()
        conn.close()
        
        if result is None:
            return jsonify({'success': False, 'message': 'Henüz yayınlanmış tablo yok'}), 404
        
        data = json.loads(result[0])
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        print(f"❌ Get published hatası: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/physicians', methods=['GET'])
def get_physicians():
    """Hekimleri getir"""
    try:
        conn = sqlite3.connect('nobet.db')
        c = conn.cursor()
        c.execute('SELECT data FROM physicians_data ORDER BY created_at DESC LIMIT 1')
        result = c.fetchone()
        conn.close()
        
        if result is None:
            return jsonify({'success': False, 'message': 'Henüz hekim kaydı yok'}), 404
        
        data = json.loads(result[0])
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        print(f"❌ Get physicians hatası: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/physicians', methods=['POST'])
def save_physicians():
    """Hekimleri kaydet"""
    try:
        data = request.json
        conn = sqlite3.connect('nobet.db')
        c = conn.cursor()
        
        c.execute('DELETE FROM physicians_data')
        c.execute('INSERT INTO physicians_data (data, created_at) VALUES (?, ?)',
                  (json.dumps(data), datetime.now()))
        conn.commit()
        conn.close()
        
        print(f"✅ {len(data)} hekim kaydedildi")
        return jsonify({'success': True, 'message': 'Hekimler kaydedildi'})
    except Exception as e:
        print(f"❌ Save physicians hatası: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/settings', methods=['GET'])
def get_settings():
    """Uygulama ayarlarını getir"""
    try:
        conn = sqlite3.connect('nobet.db')
        c = conn.cursor()
        c.execute('SELECT data FROM app_settings ORDER BY created_at DESC LIMIT 1')
        result = c.fetchone()
        conn.close()
        
        if result is None:
            return jsonify({'success': False, 'message': 'Henüz ayar kaydı yok'}), 404
        
        data = json.loads(result[0])
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        print(f"❌ Get settings hatası: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/settings', methods=['POST'])
def save_settings():
    """Uygulama ayarlarını kaydet"""
    try:
        data = request.json
        conn = sqlite3.connect('nobet.db')
        c = conn.cursor()
        
        c.execute('DELETE FROM app_settings')
        c.execute('INSERT INTO app_settings (data, created_at) VALUES (?, ?)',
                  (json.dumps(data), datetime.now()))
        conn.commit()
        conn.close()
        
        print(f"✅ Ayarlar kaydedildi")
        return jsonify({'success': True, 'message': 'Ayarlar kaydedildi'})
    except Exception as e:
        print(f"❌ Save settings hatası: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


def create_nobet_listesi():
    """
    Nöbet listesi oluştur (otomatik template, renkli)
    
    Form Data:
    - data: JSON string (physicians, schedule, year, month, daysInMonth, excuseColor)
    
    Returns: Excel dosyası
    """
    try:
        # JSON data'yı al
        import json
        data = json.loads(request.form.get('data', '{}'))
        
        # TEMPLATE'İ BACKEND'DEN YÜK (kullanıcı yüklemesin)
        template_path = os.path.join(os.path.dirname(__file__), 'templates', 'nobetprogramirenkli.xlsx')
        if not os.path.exists(template_path):
            return jsonify({'error': 'Nöbet listesi template bulunamadı!'}), 500
        
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        physicians = data.get('physicians', [])
        schedule = data.get('schedule', {})
        year = data.get('year', 2026)
        month = data.get('month', 0)
        days_in_month = data.get('daysInMonth', 31)
        excuse_color = data.get('excuseColor', 'A9A9A9')  # Varsayılan gri
        holidays = data.get('holidays', [])  # YENİ: Tatil günleri listesi
        
        month_names = ['OCAK', 'ŞUBAT', 'MART', 'NİSAN', 'MAYIS', 'HAZİRAN',
                      'TEMMUZ', 'AĞUSTOS', 'EYLÜL', 'EKİM', 'KASIM', 'ARALIK']
        
        # SATIR 9'A HEKİM BAŞLIKLARINI DOĞRUDAN YAZ (Uygulamadan gelen sırayla)
        header_row = 9
        
        for index, physician in enumerate(physicians):
            col = 3 + index  # C, D, E, F... sütunları
            cell = ws.cell(row=header_row, column=col)
            
            # Hekim kısa adını yaz
            cell.value = physician['displayName']
            
            # Hekim rengi ile boya
            hex_color = physician['color'].replace('#', '')
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        
        # Satır 10'dan başlayarak günleri doldur
        from datetime import datetime, timedelta
        start_date = datetime(year, month + 1, 1)
        
        for day in range(1, days_in_month + 1):
            row = 9 + day  # Satır 10, 11, 12...
            current_date = start_date + timedelta(days=day - 1)
            
            # A sütunu: Tarih
            ws.cell(row=row, column=1).value = current_date
            
            # B sütunu: Gün adı
            day_names = ['PAZARTESİ', 'SALI', 'ÇARŞAMBA', 'PERŞEMBE', 'CUMA', 'CUMARTESİ', 'PAZAR']
            day_name_cell = ws.cell(row=row, column=2)
            day_name_cell.value = day_names[current_date.weekday()]
            
            # Tatil kontrolü: SADECE holidays listesinde varsa gri yap
            is_holiday = day in holidays  # Frontend'den gelen tatil günleri
            
            if is_holiday:
                day_name_cell.fill = PatternFill(start_color='D1D5DB', end_color='D1D5DB', fill_type='solid')
            
            # Hekim sütunları: Nöbet saatleri + Renkler (Uygulamadaki sırayla)
            for index, physician in enumerate(physicians):
                col = 3 + index  # C, D, E, F...
                key = f"{physician['name']}-{day}"
                value = schedule.get(key, '')
                
                cell = ws.cell(row=row, column=col)
                
                # Renk ve değer belirleme
                # MAZERET TİPLERİ LİSTESİ (14 çeşit + eski formatlar)
                excuse_codes = ['YI', 'R', 'EV', 'EI', 'GO', 'MI', 'OI', 'RI', 
                               'SEI', 'KI', 'IK', 'II', 'SI', 'B', 'I', 'X']
                
                if value in ['24']:
                    # 24 saat: Hekim rengi, içi boş
                    cell.value = ''
                    hex_color = physician['color'].replace('#', '')
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                
                elif value in ['16', '8']:
                    # 16/8 saat: Hekim rengi, içinde saat
                    cell.value = value
                    hex_color = physician['color'].replace('#', '')
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                
                elif value in excuse_codes:
                    # Tüm mazeret tipleri: Gri renk (excuseColor), içi boş
                    cell.value = ''
                    hex_color = excuse_color.replace('#', '')
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                
                else:
                    # Nöbet yok: Boş, beyaz
                    cell.value = ''
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # Excel dosyasını hafızaya yaz
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'Nobet_Listesi_{month_names[month]}_{year}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/create-puantaj', methods=['POST'])
def create_puantaj():
    """
    Puantaj oluştur (otomatik template)
    
    Form Data:
    - data: JSON string (physicians, schedule, year, month, daysInMonth)
    
    Returns: Excel dosyası
    """
    try:
        # JSON data'yı al
        import json
        data = json.loads(request.form.get('data', '{}'))
        
        # TEMPLATE'İ BACKEND'DEN YÜK (kullanıcı yüklemesin)
        template_path = os.path.join(os.path.dirname(__file__), 'templates', 'puantaj.xlsx')
        if not os.path.exists(template_path):
            return jsonify({'error': 'Puantaj template bulunamadı!'}), 500
        
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Ay/Yıl güncelle
        month_names = ['OCAK', 'ŞUBAT', 'MART', 'NİSAN', 'MAYIS', 'HAZİRAN',
                      'TEMMUZ', 'AĞUSTOS', 'EYLÜL', 'EKİM', 'KASIM', 'ARALIK']
        
        year = data.get('year', 2026)
        month = data.get('month', 0)
        
        if ws['A1'].value:
            current_text = str(ws['A1'].value)
            import re
            updated_text = re.sub(r'\d{4}\s+\w+\s+AYI', 
                                f'{year} {month_names[month]} AYI', 
                                current_text)
            ws['A1'].value = updated_text
        
        physicians = data.get('physicians', [])
        schedule = data.get('schedule', {})
        days_in_month = data.get('daysInMonth', 31)
        holidays = data.get('holidays', [])  # YENİ: Tatil günleri
        
        # DEBUG
        print(f"\n=== PUANTAJ OLUŞTURMA ===")
        print(f"Physicians sayısı: {len(physicians)}")
        if physicians:
            print(f"İlk hekim: {physicians[0]}")
        print(f"Schedule sayısı: {len(schedule)}")
        
        # A1 hücresini güncelle (ay/yıl bilgisi varsa)
        if ws['A1'].value:
            current_text = str(ws['A1'].value)
            # Ay adını güncelle
            for m_name in ['OCAK', 'ŞUBAT', 'MART', 'NİSAN', 'MAYIS', 'HAZİRAN',
                          'TEMMUZ', 'AĞUSTOS', 'EYLÜL', 'EKİM', 'KASIM', 'ARALIK']:
                if m_name in current_text.upper():
                    current_text = current_text.upper().replace(m_name, month_names[month])
                    break
            ws['A1'].value = current_text
        
        # Sayfa adını da güncelle
        ws.title = month_names[month]
        
        # BAŞLIK SATIRINI BOYAMA (Satır 3 - Gün numaraları)
        # SADECE holidays listesindeki günleri gri yap
        from datetime import datetime, timedelta
        for day in range(1, days_in_month + 1):
            col = 4 + day  # E, F, G... sütunları
            header_cell = ws.cell(row=3, column=col)
            
            # Tatil kontrolü: SADECE holidays listesinde varsa
            is_holiday = day in holidays  # Frontend'den gelen tatil günleri
            
            if is_holiday:
                header_cell.fill = PatternFill(start_color='D1D5DB', end_color='D1D5DB', fill_type='solid')
        
        # HEKİMLERİ DOĞRUDAN YAZ (Satır 4'ten başlayarak)
        start_row = 4
        
        print(f"\n=== HEKİM YAZMA BAŞLIYOR ===")
        for index, physician in enumerate(physicians):
            row = start_row + index
            print(f"\nSatır {row} - Hekim {index+1}:")
            
            # A sütunu: Sıra numarası
            ws.cell(row=row, column=1).value = index + 1
            print(f"  A{row} = {index + 1}")
            
            # B sütunu: TC Kimlik No
            tc = physician.get('tc', '')
            ws.cell(row=row, column=2).value = tc
            print(f"  B{row} = {tc}")
            
            # C sütunu: Adı Soyadı
            ws.cell(row=row, column=3).value = physician['name']
            print(f"  C{row} = {physician['name']}")
            
            # D sütunu: Ünvan (Branş)
            title = physician.get('title', 'Uzm.Dr')
            ws.cell(row=row, column=4).value = title
            print(f"  D{row} = {title}")
            
            # E-AI sütunları (5-35): Nöbet saatleri (günler 1-31)
            for day in range(1, days_in_month + 1):
                key = f"{physician['name']}-{day}"
                value = schedule.get(key, '')
                
                col = 4 + day  # E=5, F=6, G=7...
                cell = ws.cell(row=row, column=col)
                
                if value:  # Sadece dolu günleri logla
                    # Değeri dönüştür
                    cell_value = None
                    
                    # MAZERET TİPLERİ (14 çeşit)
                    excuse_map = {
                        'YI': 'Yİ',   # Yıllık İzin
                        'R': 'R',     # Raporlu
                        'EV': 'EV',   # Evlilik İzni
                        'EI': 'Eİ',   # Eğitim İzni
                        'GO': 'GÖ',   # Görevlendirme
                        'MI': 'Mİ',   # Mazeret İzni
                        'OI': 'Öİ',   # Ölüm İzni
                        'RI': 'Rİ',   # Refakat İzni
                        'SEI': 'SEİ', # Sendika İzni
                        'KI': 'Kİ',   # Kongre İzni
                        'IK': 'İK',   # İş Kazası
                        'II': 'İİ',   # İdari İzin
                        'SI': 'Sİ',   # Sevk İzin
                        'B': None,    # Boşluk - boş bırak
                        # Eski formatlar (geriye dönük uyumluluk)
                        'I': 'Yİ',
                        'X': 'R'
                    }
                    
                    if value in excuse_map:
                        cell_value = excuse_map[value]
                    elif value in ['24', '16', '8']:
                        cell_value = value  # Nöbet saati
                    
                    # Hücreye yaz
                    cell.value = cell_value
                    if cell_value:
                        print(f"  Gün {day}: {cell_value}")
        
        print(f"\n=== YAZMA BİTTİ ===")
        print(f"Toplam {len(physicians)} hekim yazıldı")
        
        # Excel dosyasını hafızaya yaz
        print(f"\n=== EXCEL KAYIT ===")
        print(f"Worksheet title: {ws.title}")
        print(f"Max row: {ws.max_row}")
        print(f"A4 değeri: {ws.cell(4, 1).value}")
        print(f"B4 değeri: {ws.cell(4, 2).value}")
        print(f"C4 değeri: {ws.cell(4, 3).value}")
        print(f"D4 değeri: {ws.cell(4, 4).value}")
        
        output = io.BytesIO()
        
        try:
            wb.save(output)
            print(f"✅ Workbook kaydedildi")
        except Exception as save_error:
            print(f"❌ Save hatası: {save_error}")
            raise
        
        output.seek(0)
        print(f"Output size: {output.getbuffer().nbytes} bytes")
        
        filename = f'Puantaj_{month_names[month]}_{year}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload-template', methods=['POST'])
def upload_template():
    """Template dosyasını yükle ve hekim listesini döndür"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Dosya bulunamadı'}), 400
        
        file = request.files['file']
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        doctors = []
        for row in range(4, 26):
            doctor_cell = ws.cell(row=row, column=3)  # C sütunu
            if doctor_cell.value and str(doctor_cell.value).strip():
                doctors.append({
                    'name': str(doctor_cell.value).strip(),
                    'row': row
                })
        
        return jsonify({
            'doctors': doctors,
            'count': len(doctors)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/create-nobet-isimleri', methods=['POST'])
def create_nobet_isimleri():
    """
    Nöbet isimleri listesi oluştur (otomatik template)
    
    Form Data:
    - data: JSON string (physicians, schedule, year, month, daysInMonth)
    
    Returns: Excel dosyası
    """
    try:
        # JSON data'yı al
        import json
        data = json.loads(request.form.get('data', '{}'))
        
        physicians = data.get('physicians', [])
        schedule = data.get('schedule', {})
        year = data.get('year', 2026)
        month = data.get('month', 0)
        days_in_month = data.get('daysInMonth', 31)
        holidays = data.get('holidays', [])  # YENİ: Tatil günleri
        
        print(f"\n=== NÖBET İSİMLERİ DEBUG ===")
        print(f"Physicians sayısı: {len(physicians)}")
        for i, p in enumerate(physicians, 1):
            print(f"  {i}. {p.get('name')} - {p.get('title')}")
        print(f"Schedule keys: {len(schedule)}")
        print(f"Year: {year}, Month: {month}, Days: {days_in_month}")
        print(f"Holidays: {holidays}")
        
        # TEMPLATE'İ BACKEND'DEN YÜK
        template_path = os.path.join(os.path.dirname(__file__), 'templates', 'nobetisimler_template.xlsx')
        if not os.path.exists(template_path):
            return jsonify({'error': 'Nöbet isimleri template bulunamadı!'}), 500
        
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        month_names = ['OCAK', 'ŞUBAT', 'MART', 'NİSAN', 'MAYIS', 'HAZİRAN',
                      'TEMMUZ', 'AĞUSTOS', 'EYLÜL', 'EKİM', 'KASIM', 'ARALIK']
        
        # Başlık güncelle (A5)
        ws['A5'].value = f"  {year} {month_names[month]} AYI NÖBET LİSTESİ"
        
        # Satır 7'den itibaren günleri doldur
        from datetime import datetime, timedelta
        start_date = datetime(year, month + 1, 1)
        
        for day in range(1, days_in_month + 1):
            row = 6 + day  # Satır 7, 8, 9...
            current_date = start_date + timedelta(days=day - 1)
            
            # A sütunu: Tarih
            ws.cell(row=row, column=1).value = current_date
            
            # B sütunu: Gün adı
            day_names = ['PAZARTESİ', 'SALI', 'ÇARŞAMBA', 'PERŞEMBE', 'CUMA', 'CUMARTESİ', 'PAZAR']
            day_name_cell = ws.cell(row=row, column=2)
            day_name_cell.value = day_names[current_date.weekday()]
            
            # Tatil kontrolü: SADECE holidays listesinde varsa
            is_holiday = day in holidays
            
            if is_holiday:
                day_name_cell.fill = PatternFill(start_color='D1D5DB', end_color='D1D5DB', fill_type='solid')
            
            # O gün nöbet tutan hekimleri bul
            mesai_doctors = []  # 8 saat
            nobet_24_doctors = []  # 24 saat
            nobet_16_doctors = []  # 16 saat
            
            for physician in physicians:
                key = f"{physician['name']}-{day}"
                value = schedule.get(key, '')
                
                if value == '8':
                    # Ünvan ekle + TAM AD SOYAD (name, displayName değil)
                    title = 'UZM.DR.' if physician.get('title') == 'Uzm.Dr' else 'DR.'
                    full_name = f"{title}{physician['name']}"  # TAM AD SOYAD
                    mesai_doctors.append(full_name)
                elif value == '24':
                    title = 'UZM.DR.' if physician.get('title') == 'Uzm.Dr' else 'DR.'
                    full_name = f"{title}{physician['name']}"  # TAM AD SOYAD
                    nobet_24_doctors.append(full_name)
                elif value == '16':
                    title = 'UZM.DR.' if physician.get('title') == 'Uzm.Dr' else 'DR.'
                    full_name = f"{title}{physician['name']}"  # TAM AD SOYAD
                    nobet_16_doctors.append(f"{full_name}(16)")
            
            # C sütunu: MESAİ (8 saat) - + ile birleştir
            if mesai_doctors:
                ws.cell(row=row, column=3).value = '+'.join(mesai_doctors)
            
            # D, E, F sütunları: NÖBET (24+16 saat)
            # 24 ve 16 saatleri birleştir
            all_nobet = nobet_24_doctors + nobet_16_doctors
            
            # İlk 3 tanesi D, E, F'ye
            for i, doctor_name in enumerate(all_nobet[:3]):
                ws.cell(row=row, column=4 + i).value = doctor_name
            
            # 3'ten fazlaysa, son sütuna + ile ekle
            if len(all_nobet) > 3:
                last_cell = ws.cell(row=row, column=6)  # F sütunu
                existing = last_cell.value or ''
                additional = '+'.join(all_nobet[3:])
                last_cell.value = f"{existing}+{additional}" if existing else additional
        
        # G SÜTUNUNU DOLDUR - TÜM HEKİMLER (Her Excel satırına 1 hekim)
        # G7: 1.Hekim1, G8: 2.Hekim2, G9: 3.Hekim3...
        for idx, physician in enumerate(physicians, 1):
            row = 6 + idx  # G7, G8, G9...
            
            # Ünvan belirleme
            if physician.get('title') == 'Uzm.Dr':
                title = 'ACİL UZM.DR.'  # Uzman doktorlara "ACİL" ön eki
            else:
                title = 'DR.'
            
            full_name = f"{title}{physician['name']}"
            ws.cell(row=row, column=7).value = f"{idx}.{full_name}"
        
        print(f"\n=== G SÜTUNU TAMAMLANDI ===")
        print(f"G7-G{6+len(physicians)}: Tüm hekimler listelenmiş")
        print(f"Toplam {len(physicians)} hekim")
        
        # Excel dosyasını hafızaya yaz
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, download_name=f'Nobet_Isimleri_{month_names[month]}_{year}.xlsx', as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    except Exception as e:
        print(f"❌ HATA: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    # Production: Render.com PORT environment variable kullanır
    # Development: 5000 portu kullanır
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
